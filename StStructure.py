import streamlit as st
import pandas as pd
import math
import sympy as sp
from openpyxl import load_workbook


st.title('App Cálculo estrutural 2D')

def Prop():
    Propriedades = []
    MI = []
    ME =[]
    AR = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaProp.cell(i,1).value) != None:
        MI.append(AbaProp.cell(i,1).value)
        ME.append(AbaProp.cell(i,2).value)
        AR.append(AbaProp.cell(i,3).value)
        i +=1
    Propriedades.append(MI)
    Propriedades.append(ME)
    Propriedades.append(AR)
    return Propriedades
    
def Nosf():
    Nos = []
    x = []
    y = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaNos.cell(i,1).value) != None:
        No = []
        No.append(AbaNos.cell(i,1).value)
        No.append(AbaNos.cell(i,2).value)
        Nos.append(No)
        i +=1
    return Nos


def barras():
    barras = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaBarras.cell(i,1).value) != None:
        barra = []
        barra.append(AbaBarras.cell(i,1).value)
        barra.append(AbaBarras.cell(i,2).value)
        barra.append(AbaBarras.cell(i,3).value)
        barra.append(AbaBarras.cell(i,4).value)
        barra.append(AbaBarras.cell(i,5).value)
        barras.append(barra)
        i +=1
    return barras

def Forcas():
    forcas = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaForca.cell(i,1).value) != None:
        forca = []
        forca.append(AbaForca.cell(i,1).value)
        forca.append(AbaForca.cell(i,2).value)
        forca.append(AbaForca.cell(i,3).value)
        forca.append(AbaForca.cell(i,4).value)
        forcas.append(forca)
        i +=1
    return forcas

def Apoios():
    apoios = []
    filename = 'Valores.xlsx'
    planilha = load_workbook(filename)
    i=1
    while (AbaApoio.cell(i,1).value) != None:
        apoio = []
        apoio.append(AbaApoio.cell(i,1).value)
        apoio.append(AbaApoio.cell(i,2).value)
        apoio.append(AbaApoio.cell(i,3).value)
        apoio.append(AbaApoio.cell(i,4).value)
        apoios.append(apoio)
        i +=1
    return apoios

filename = 'Valores.xlsx'
planilha = load_workbook(filename)

AbaProp = planilha['AbaProp']
AbaNos = planilha['AbaNos']
AbaBarras = planilha['AbaBarras']
AbaForca = planilha['AbaForca']
AbaApoio = planilha['AbaApoio']

st.sidebar.header('Propriedades')
I = st.sidebar.number_input('Momento de Inércia:')
E = st.sidebar.number_input('Módulo de Elasticidade:')          
A = st.sidebar.number_input('Área:')

if st.sidebar.button('Confirmar'):
    i=1
    while (AbaProp.cell(i,1).value) != None:
        i +=1
    AbaProp.cell(i,1).value = I
    AbaProp.cell(i,2).value = E
    AbaProp.cell(i,3).value = A
    planilha.save(filename)

st.sidebar.header('Nós')
x = st.sidebar.number_input('Insira coordenada x')
y = st.sidebar.number_input('Insira coordenada y')

if st.sidebar.button('Confirmar Nó'):
    i=1
    while (AbaNos.cell(i,1).value) != None:
        i +=1
    AbaNos.cell(i,1).value = x
    AbaNos.cell(i,2).value = y
    planilha.save(filename)


nos = Nosf()
Propriedades = Prop()
nprop = len(Propriedades[0])
if (nprop == 0):
    nprop=1


if st.sidebar.button('Mostrar Propriedades'):
    st.write(Propriedades)
if st.sidebar.button('Mostrar Nos'):
    st.write(nos)

st.sidebar.header('Barras')
no1 = st.sidebar.number_input('Insira o nó1',value=0)
no2 = st.sidebar.number_input('Insira o nó2',value=0)
prop = st.sidebar.slider('Selecione um padrão de barra',0,nprop,0)

if st.sidebar.button('Confirmar barra'):
    i=1
    while (AbaBarras.cell(i,1).value) != None:
        i +=1
    AbaBarras.cell(i,1).value = no1
    AbaBarras.cell(i,2).value = no2
    AbaBarras.cell(i,3).value = Propriedades[0][prop]
    AbaBarras.cell(i,4).value = Propriedades[1][prop]
    AbaBarras.cell(i,5).value = Propriedades[2][prop]
    planilha.save(filename)
    
barras = barras()

if st.sidebar.button('Mostrar Barras'):
    st.write(barras)

st.sidebar.header('Forças')

no = st.sidebar.number_input('Insira o nó',value=0)
Fx = st.sidebar.number_input('Insira a decomposição x da Força')
Fy = st.sidebar.number_input('Insira a decomposição y da Força')
M = st.sidebar.number_input('Insira o momento')

if st.sidebar.button('Confirmar Força'):
    i=1
    while (AbaForca.cell(i,1).value) != None:
        i +=1
    AbaForca.cell(i,1).value = no
    AbaForca.cell(i,2).value = Fx
    AbaForca.cell(i,3).value = Fy
    AbaForca.cell(i,4).value = M
    planilha.save(filename)

forcas = Forcas()

if st.sidebar.button('Mostrar Forças'):
    st.write(forcas)

st.sidebar.header('Apoios')

noap = st.sidebar.number_input('Insira o nó do apoio',value=0)
restricaox = st.sidebar.checkbox ('O movimento é restrito em x?')
restricaoy = st.sidebar.checkbox ('O movimento é restrito em y?')
restricaoz = st.sidebar.checkbox ('O movimento é restrito em z?')

if st.sidebar.button('Confirmar Apoio'):
    i=1
    while (AbaApoio.cell(i,1).value) != None:
        i +=1
    AbaApoio.cell(i,1).value = noap
    if restricaox:
        AbaApoio.cell(i,2).value = True
    else:
        AbaApoio.cell(i,2).value = False
    if restricaoy:
        AbaApoio.cell(i,3).value = True
    else:
        AbaApoio.cell(i,3).value = False
    if restricaoz:
        AbaApoio.cell(i,4).value = True
    else:
        AbaApoio.cell(i,4).value = False
    planilha.save(filename)

apoios = Apoios()

if st.sidebar.button('Mostrar Apoios'):
    st.write(apoios)

if st.sidebar.button('Limpar arquivo'):
    AbaProp.delete_cols(3)
    AbaProp.delete_cols(2)
    AbaProp.delete_cols(1)
    AbaNos.delete_cols(2)
    AbaNos.delete_cols(1)
    AbaBarras.delete_cols(5)
    AbaBarras.delete_cols(4)
    AbaBarras.delete_cols(3)
    AbaBarras.delete_cols(2)
    AbaBarras.delete_cols(1)
    AbaForca.delete_cols(4)
    AbaForca.delete_cols(3)
    AbaForca.delete_cols(2)
    AbaForca.delete_cols(1)
    AbaApoio.delete_cols(4)
    AbaApoio.delete_cols(3)
    AbaApoio.delete_cols(2)
    AbaApoio.delete_cols(1)
    planilha.save(filename)


QTDE_BARRAS = len(barras)

# Cálculo do comprimento e ângulo de cada barra
for barra in barras:
    no1 = nos[barra[0]]
    no2 = nos[barra[1]]
    barra.append(((no2[0] - no1[0]) ** 2 + (no2[1] - no1[1]) ** 2) ** 0.5)
    barra.append(math.atan2(no2[1] - no1[1], no2[0] - no1[0]))

# Quantidade de nós e graus de liberdade
QTDE_NOS = len(nos)
GDL = 3 * QTDE_NOS

# Definição dos símbolos
A, E, L, I, theta = sp.symbols(["A", "E", "L", "I", "theta"])

# Definição das variáveis para a matriz de rigidez
COS = sp.cos(theta)
SIN = sp.sin(theta)
T = sp.Matrix(
    [
        [COS, -SIN, 0, 0, 0, 0],
        [SIN, COS, 0, 0, 0, 0],
        [0, 0, 1, 0, 0, 0],
        [0, 0, 0, COS, -SIN, 0],
        [0, 0, 0, SIN, COS, 0],
        [0, 0, 0, 0, 0, 1],
    ]
).T

Ke_ = sp.Matrix(
    [
        [E * A / L, 0, 0, -E * A / L, 0, 0],
        [
            0,
            12 * E * I / L**3,
            6 * E * I / L**2,
            0,
            -12 * E * I / L**3,
            6 * E * I / L**2,
        ],
        [0, 6 * E * I / L**2, 4 * E * I / L, 0, -6 * E * I / L**2, 2 * E * I / L],
        [-E * A / L, 0, 0, E * A / L, 0, 0],
        [
            0,
            -12 * E * I / L**3,
            -6 * E * I / L**2,
            0,
            12 * E * I / L**3,
            -6 * E * I / L**2,
        ],
        [0, 6 * E * I / L**2, 2 * E * I / L, 0, -6 * E * I / L**2, 4 * E * I / L],
    ]
)

Ke = T.T * Ke_ * T

# Montagem da matriz de rigidez
lista_Ke = []
for i in range(QTDE_BARRAS):
    lista_Ke.append(
        Ke.subs(
            [
                (I, barras[i][2]),
                (E, barras[i][3]),
                (A, barras[i][4]),
                (L, barras[i][5]),
                (theta, barras[i][6]),
            ]
        )
    )

# Montagem da matriz de rigidez global
K = sp.zeros(GDL, GDL)
for i in range(QTDE_BARRAS):
    no1 = barras[i][0]
    no2 = barras[i][1]
    indices = [3 * no1, 3 * no1 + 1, 3 * no1 + 2, 3 * no2, 3 * no2 + 1, 3 * no2 + 2]
    for j in range(6):
        for k in range(6):
            K[indices[j], indices[k]] += lista_Ke[i][j, k]

# Montagem do vetor de forças
S = [0] * GDL
for i in range(len(forcas)):
    no = forcas[i][0]
    ForcaX = forcas[i][1]
    ForcaY = forcas[i][2]
    momento = forcas[i][3]
    if type(ForcaX) == str:
        S[3* no] = sp.symbols(["S"+str(3*no+1)])
    else:
        S[3 * no] += ForcaX
    if type(ForcaY) ==str:
        S[3 * no + 1] = sp.symbols(["S"+str(3*no+2)])
    else:
        S[3 * no + 1] += ForcaY
    if type(momento) == str:
        S[3 * no + 2] = sp.symbols(["S"+str(3*no+3)])
    else: 
        S[3 * no + 2] += momento
S = sp.Matrix(S)

# Montagem do vetor de deslocamentos
q = sp.symbols(["q" + str(i) for i in range(1, GDL + 1)])
for i in range(len(apoios)):
    no = apoios[i][0]
    if apoios[i][1]:
        q[3 * no] = 0
        S[3 * no] = sp.symbols("S" + str(3 * no + 1))
    if apoios[i][2]:
        q[3 * no + 1] = 0
        S[3 * no + 1] = sp.symbols("S" + str(3 * no + 2))
    if apoios[i][3]:
        q[3 * no + 2] = 0
        S[3 * no + 2] = sp.symbols("S" + str(3 * no + 3))
q = sp.Matrix(q)

# Separando variáveis a serem resolvidas
variaveis_sistema = [i for i in q if isinstance(i, sp.Symbol)] + [
    i for i in S if isinstance(i, sp.Symbol)
]

# Resolução do sistema
sistema = K * q - S



if st.button('Calcular'): 
    resolucao = sp.solve(sistema, variaveis_sistema)
    resultado = str(resolucao)
    st.write(resultado)














    


